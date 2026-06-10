"""
================================================================================
MSCI Annual Update Factual Process Automation — STREAMLIT WEB APP
================================================================================
Version     : 13.0.0-web
Based on    : annual_update_processor_v13.txt (console version)

What changed vs the console version
-----------------------------------
1. No hardcoded Windows folder paths. The user uploads the Extraction CSV and
   the DVV Merged XLSX directly in the browser. The 3 templates are either
   bundled in a local  ./templates  folder or uploaded each run.
2. input()/main() console flow replaced by a Streamlit UI (uploaders, text box,
   Run button, download buttons).
3. Outputs are generated in-memory (BytesIO) and offered as individual
   download buttons plus a single "Download all (ZIP)" button. Nothing is
   written to disk on the server beyond a short-lived temp folder that is
   deleted at the end of every run.
4. Logging is captured into the page (expandable log panel) instead of stdout.

The CORE processing logic (header standardisation, extraction load, DVV load,
template population, DVV overrides, formatting, validation) is IDENTICAL to
v13 — only the I/O layer around it changed, so output files are byte-for-byte
equivalent to the desktop script.

DVV column mapping note
-----------------------
The DVV file gained a new column "UUID4" at position A, shifting everything
one column right. v13's mapping already reflects this:
    G  -> DATALIB_TAG      AH -> CORRECT_VALUE      AM -> SERIALID

Run locally
-----------
    pip install streamlit pandas openpyxl
    streamlit run annual_update_app.py

Requirements: Python 3.11+ | streamlit | pandas | openpyxl
================================================================================
"""

import io
import re
import zipfile
import logging
import datetime
import tempfile
import traceback
import shutil
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import streamlit as st


# ==============================================================================
# SECTION 1: CONFIGURATION
# Only the input/output *folder* paths from the console version were removed
# (the web app uses uploads + a temp folder instead). Everything else — column
# letters, template filenames, formatting, exclusions — is unchanged from v13.
# ==============================================================================

CONFIG = {
    # ── Template file names (exact filenames expected on disk / in uploads) ───
    # Scalar and Series1 are combined into one file (Scalar_Series1).
    "TEMPLATES": {
        "Position":       "Position_Tab_Split.xlsx",
        "Scalar_Series1": "Scalar & Series 1_DP.xlsx",
        "Series2":        "Series 2- Delete blank cell_DP.xlsx",
    },

    # ── Output file name suffixes →  IssuerID_<suffix>.xlsx ───────────────────
    "OUTPUT_NAMES": {
        "Position":       "Position_Tab_Split",
        "Scalar_Series1": "Scalar & Series 1_DP",
        "Series2":        "Series 2- Delete blank cell_DP",
    },

    # ── Name of the scalar (issuer-level) sheet inside Scalar_Series1 ─────────
    "SCALAR_SHEET_NAME": "Scalar",

    # ── DVV TAB exclusions (rows on these tabs are dropped before overrides) ──
    "DVV_EXCLUDE_TABS": {
        "Director Attributes",
        "Director Data - Board",
        "Positions",
        "Individual",
    },

    # ── DVV column letters (verified against the latest DVV Merged file) ──────
    # A new "UUID4" column at position A shifted these one column right vs v1.
    "DVV_DATALIB_COL":    "G",   # Column G  → DATALIB_TAG   (e.g. DIRGENDER)
    "DVV_CORRECTVAL_COL": "AH",  # Column AH → CORRECT_VALUE
    "DVV_SERIES_COL":     "AM",  # Column AM → SERIALID

    # ── Extraction Series ID column index (0-based) in the RAW CSV ────────────
    # Column AD (index 29) = 'Series ID' (numeric IDs that match DVV SERIALID).
    "EXTRACTION_SERIES_COL_INDEX": 29,

    # ── Field names ───────────────────────────────────────────────────────────
    "ISSUER_ID_FIELD": "DMX_ISSUER_ID",

    # ── Formatting ────────────────────────────────────────────────────────────
    "HEADER_FONT_NAME":    "Times New Roman",
    "HEADER_FONT_SIZE":    12,
    "HEADER_FONT_COLOR":   "FFFFFF",
    "HEADER_FILL_COLOR":   "0070C0",
    "DATA_FONT_NAME":      "Times New Roman",
    "DATA_FONT_SIZE":      12,
    "DVV_HIGHLIGHT_COLOR": "E6FDCF",   # Light green for DVV-updated cells
}

# Where to look for bundled template files (so users don't re-upload them).
# Place the 3 template .xlsx files in a folder named "templates" next to this
# script and they will be picked up automatically.
TEMPLATES_DIR = Path(__file__).parent / "templates" if "__file__" in globals() \
    else Path("templates")


# ==============================================================================
# SECTION 2: EXTRACTION CSV PROCESSING  (unchanged from v13)
# ==============================================================================

def standardize_header(raw_header: str) -> str:
    """
    Convert a CSV header to its Data Lib identifier.
        DMX_ISSUER_ID                      -> DMX_ISSUER_ID
        Audit Board Member (REL_AUDIT_BOARD) -> REL_AUDIT_BOARD
    Rule: extract text inside the LAST pair of parentheses; else return as-is.
    """
    raw = str(raw_header).strip()
    match = re.search(r"\(([^)]+)\)\s*$", raw)
    if match:
        return match.group(1).strip()
    return raw


def load_extraction(csv_path: Path, issuer_id: str,
                    logger: logging.Logger) -> pd.DataFrame:
    """Load extraction CSV and standardise headers. All values read as strings."""
    logger.info(f"Loading extraction CSV: {csv_path.name}")

    try:
        df = pd.read_csv(csv_path, dtype=str, encoding="utf-8-sig")
    except UnicodeDecodeError:
        df = pd.read_csv(csv_path, dtype=str, encoding="latin-1")

    if df.empty:
        raise ValueError(f"Extraction CSV is empty: {csv_path.name}")

    df.columns = [standardize_header(c) for c in df.columns]

    if len(df.columns) > CONFIG["EXTRACTION_SERIES_COL_INDEX"]:
        series_col_name = df.columns[CONFIG["EXTRACTION_SERIES_COL_INDEX"]]
        logger.info(
            f"Series ID column (index {CONFIG['EXTRACTION_SERIES_COL_INDEX']}) "
            f"= '{series_col_name}'"
        )
    else:
        logger.warning(
            f"Extraction CSV has fewer than "
            f"{CONFIG['EXTRACTION_SERIES_COL_INDEX']+1} columns — "
            "Series ID column cannot be read."
        )

    logger.info(
        f"Extraction loaded: {len(df)} rows, "
        f"{len(df.columns)} columns after header standardisation"
    )
    return df


def build_datalib_to_series_map(ext_df: pd.DataFrame) -> dict:
    """
    Map datalib_code -> Series ID column index. The extraction CSV has multiple
    'Series ID' columns (one per section); each precedes its section's data
    columns. The correct Series ID for a code is the nearest one BEFORE it.
    """
    cols = list(ext_df.columns)
    series_id_positions = [
        i for i, c in enumerate(cols)
        if c == "Series ID" or re.match(r"^Series ID(\.\d+)?$", c)
    ]

    datalib_to_series_col: dict = {}
    for col_idx, col_name in enumerate(cols):
        if col_name in ("Series ID",) or re.match(r"^Series ID(\.\d+)?$", col_name):
            continue
        before = [p for p in series_id_positions if p < col_idx]
        if before:
            datalib_to_series_col[col_name] = before[-1]

    return datalib_to_series_col


def get_series_id_for_datalib(ext_row: pd.Series, ext_df: pd.DataFrame,
                              datalib_code: str,
                              datalib_to_series_map: dict):
    """Return the Series ID value for a specific data lib code in a row."""
    series_col_idx = datalib_to_series_map.get(datalib_code)
    if series_col_idx is None:
        return None
    col_name = ext_df.columns[series_col_idx]
    val = ext_row.get(col_name)
    if pd.notna(val) and str(val).strip() not in ("", "nan"):
        return str(val).strip()
    return None


def row_has_data(ext_row: pd.Series, template_headers: list) -> bool:
    """True if the row has at least one non-blank value among template headers."""
    skip = {CONFIG["ISSUER_ID_FIELD"], "serial_id"}
    for h in template_headers:
        if h in skip:
            continue
        val = ext_row.get(h)
        if pd.notna(val) and str(val).strip() not in ("", "nan"):
            return True
    return False


# ==============================================================================
# SECTION 3: DVV FILE PROCESSING  (unchanged from v13)
# ==============================================================================

def load_dvv(dvv_path: Path, logger: logging.Logger) -> pd.DataFrame:
    """
    Load DVV Merged XLSX. Keep only rows with a non-blank Correct Value (AH).
        G  (idx 6)  → DATALIB_TAG
        AH (idx 33) → CORRECT_VALUE
        AM (idx 38) → SERIALID
    Then drop excluded TAB rows.
    """
    logger.info(f"Loading DVV file: {dvv_path.name}")

    def col_letter_to_index(letter: str) -> int:
        letter = letter.upper()
        result = 0
        for ch in letter:
            result = result * 26 + (ord(ch) - ord('A') + 1)
        return result - 1

    try:
        dvv_df = pd.read_excel(
            dvv_path, sheet_name=0, dtype=str, header=0, engine="openpyxl"
        )
    except Exception as e:
        raise RuntimeError(f"Cannot read DVV file '{dvv_path.name}': {e}")

    if dvv_df.empty:
        raise ValueError(f"DVV file is empty: {dvv_path.name}")

    datalib_idx    = col_letter_to_index(CONFIG["DVV_DATALIB_COL"])
    correctval_idx = col_letter_to_index(CONFIG["DVV_CORRECTVAL_COL"])
    series_idx     = col_letter_to_index(CONFIG["DVV_SERIES_COL"])

    max_needed = max(datalib_idx, correctval_idx, series_idx)
    if dvv_df.shape[1] <= max_needed:
        raise ValueError(
            f"DVV file has only {dvv_df.shape[1]} columns; "
            f"need at least {max_needed + 1} "
            f"(up to column {CONFIG['DVV_SERIES_COL']})."
        )

    cols = dvv_df.columns.tolist()
    logger.info(
        f"DVV columns → DataLib='{cols[datalib_idx]}' "
        f"CorrectVal='{cols[correctval_idx]}' "
        f"SeriesID='{cols[series_idx]}'"
    )

    dvv_df = dvv_df.rename(columns={
        cols[datalib_idx]:    "_DVV_DATALIB",
        cols[correctval_idx]: "_DVV_CORRECT_VALUE",
        cols[series_idx]:     "_DVV_SERIES_ID",
    })

    before = len(dvv_df)
    dvv_df = dvv_df[
        dvv_df["_DVV_CORRECT_VALUE"].notna() &
        (dvv_df["_DVV_CORRECT_VALUE"].str.strip() != "")
    ].copy()
    logger.info(
        f"DVV loaded: {before} total rows, {len(dvv_df)} rows with Correct Values"
    )

    tab_col = "TAB" if "TAB" in dvv_df.columns else None
    if tab_col and CONFIG["DVV_EXCLUDE_TABS"]:
        before_tab = len(dvv_df)
        dvv_df = dvv_df[~dvv_df[tab_col].isin(CONFIG["DVV_EXCLUDE_TABS"])].copy()
        logger.info(
            f"DVV TAB filter: excluded {before_tab - len(dvv_df)} rows "
            f"from tabs: {CONFIG['DVV_EXCLUDE_TABS']}. Remaining: {len(dvv_df)} rows."
        )

    return dvv_df


# ==============================================================================
# SECTION 4: TEMPLATE POPULATION  (unchanged from v13)
# ==============================================================================

def populate_template(template_path: Path, template_key: str,
                      ext_df: pd.DataFrame, issuer_id: str,
                      logger: logging.Logger):
    """Populate all sheets in a template workbook from the extraction DataFrame."""
    logger.info(f"Populating '{template_key}': {template_path.name}")
    wb = load_workbook(template_path)
    sheet_meta = {}

    datalib_to_series_map = build_datalib_to_series_map(ext_df)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        headers = [cell.value for cell in ws[1] if cell.value is not None]
        if not headers:
            logger.warning(f"  Sheet '{sheet_name}' has no headers — skipping.")
            continue

        header_col_map = {
            cell.value: col_idx
            for col_idx, cell in enumerate(ws[1], start=1)
            if cell.value is not None
        }

        is_scalar = (sheet_name == CONFIG["SCALAR_SHEET_NAME"])
        data_start_row = 2
        populated_rows = 0

        # ── SCALAR: one issuer row ─────────────────────────────────────────
        if is_scalar:
            if CONFIG["ISSUER_ID_FIELD"] in ext_df.columns:
                issuer_rows = ext_df[ext_df[CONFIG["ISSUER_ID_FIELD"]] == issuer_id]
                if issuer_rows.empty:
                    issuer_rows = ext_df
            else:
                issuer_rows = ext_df

            if issuer_rows.empty:
                logger.warning(
                    f"  No rows for issuer {issuer_id}. "
                    f"Scalar sheet '{sheet_name}' left empty."
                )
                continue

            ext_row   = issuer_rows.iloc[0]
            write_row = data_start_row

            for tmpl_header, col_idx in header_col_map.items():
                if tmpl_header == CONFIG["ISSUER_ID_FIELD"]:
                    ws.cell(row=write_row, column=col_idx).value = issuer_id
                elif tmpl_header in ext_df.columns:
                    val = ext_row.get(tmpl_header)
                    ws.cell(row=write_row, column=col_idx).value = (
                        None if (pd.isna(val) or str(val).strip() in ("", "nan"))
                        else str(val).strip()
                    )
            populated_rows = 1

        # ── POSITION / SERIES: one row per extraction row ──────────────────
        else:
            sheet_series_col_idx = None
            sheet_section_cols   = []

            for h in headers:
                if h in (CONFIG["ISSUER_ID_FIELD"], "serial_id"):
                    continue
                idx = datalib_to_series_map.get(h)
                if idx is not None and sheet_series_col_idx is None:
                    sheet_series_col_idx = idx

            if sheet_series_col_idx is not None:
                all_std = list(ext_df.columns)
                series_positions = [
                    i for i, c in enumerate(all_std)
                    if c == "Series ID" or re.match(r"^Series ID(\.\d+)?$", c)
                ]
                next_series = next(
                    (p for p in series_positions if p > sheet_series_col_idx),
                    len(all_std)
                )
                sheet_section_cols = [
                    all_std[i]
                    for i in range(sheet_series_col_idx + 1, next_series)
                    if all_std[i] in headers
                ]

            for _, ext_row in ext_df.iterrows():
                # Gate 1: section Series ID must be non-blank
                if sheet_series_col_idx is not None:
                    sid_col_name = ext_df.columns[sheet_series_col_idx]
                    sid_val = ext_row.get(sid_col_name)
                    if pd.isna(sid_val) or str(sid_val).strip() in ("", "nan"):
                        continue
                    section_series_id = str(sid_val).strip()
                else:
                    section_series_id = None

                # Gate 2: at least one section DATA column must be non-blank
                if sheet_section_cols:
                    section_has_data = any(
                        pd.notna(ext_row.get(c)) and
                        str(ext_row.get(c, "")).strip() not in ("", "nan")
                        for c in sheet_section_cols
                    )
                    if not section_has_data:
                        continue
                elif not row_has_data(ext_row, headers):
                    continue

                write_row = data_start_row + populated_rows
                for tmpl_header, col_idx in header_col_map.items():
                    if tmpl_header == CONFIG["ISSUER_ID_FIELD"]:
                        ws.cell(row=write_row, column=col_idx).value = issuer_id
                    elif tmpl_header == "serial_id":
                        ws.cell(row=write_row, column=col_idx).value = section_series_id
                    elif tmpl_header in ext_df.columns:
                        val = ext_row.get(tmpl_header)
                        ws.cell(row=write_row, column=col_idx).value = (
                            None if (pd.isna(val) or str(val).strip() in ("", "nan"))
                            else str(val).strip()
                        )
                populated_rows += 1

        sheet_meta[sheet_name] = {
            "headers":        headers,
            "header_col_map": header_col_map,
            "data_start_row": data_start_row,
            "populated_rows": populated_rows,
        }
        logger.debug(f"  Sheet '{sheet_name}': {populated_rows} row(s) written.")

    return wb, sheet_meta


# ==============================================================================
# SECTION 5: DVV OVERRIDE PROCESSING  (unchanged from v13)
# ==============================================================================

def apply_dvv_overrides(wb, sheet_meta, dvv_df, issuer_id, template_key,
                        logger: logging.Logger) -> int:
    """Apply DVV Correct Value overrides to the populated workbook."""
    dvv_fill = PatternFill(
        "solid",
        start_color=CONFIG["DVV_HIGHLIGHT_COLOR"],
        end_color=CONFIG["DVV_HIGHLIGHT_COLOR"],
    )
    total_overrides = 0

    for sheet_name in wb.sheetnames:
        if sheet_name not in sheet_meta:
            continue

        ws             = wb[sheet_name]
        meta           = sheet_meta[sheet_name]
        header_col_map = meta["header_col_map"]
        data_start_row = meta["data_start_row"]

        is_scalar     = (sheet_name == CONFIG["SCALAR_SHEET_NAME"])
        serial_id_col = header_col_map.get("serial_id")
        issuer_id_col = header_col_map.get(CONFIG["ISSUER_ID_FIELD"])

        row_map: dict = {}
        if not is_scalar and serial_id_col:
            for r in range(data_start_row, data_start_row + meta["populated_rows"]):
                sid = ws.cell(row=r, column=serial_id_col).value
                if sid is not None:
                    row_map[str(sid).strip()] = r

        sheet_overrides = 0

        for _, dvv_row in dvv_df.iterrows():
            datalib     = str(dvv_row.get("_DVV_DATALIB", "")).strip()
            correct_val = dvv_row.get("_DVV_CORRECT_VALUE")
            dvv_series  = str(dvv_row.get("_DVV_SERIES_ID", "")).strip()
            if dvv_series.lower() in ("nan", "none", ""):
                dvv_series = ""

            if not datalib or datalib not in header_col_map:
                continue

            target_col = header_col_map[datalib]

            if is_scalar:
                ws.cell(row=data_start_row, column=target_col).value = correct_val
                ws.cell(row=data_start_row, column=target_col).fill  = dvv_fill

            elif dvv_series and serial_id_col:
                if dvv_series in row_map:
                    row_num = row_map[dvv_series]
                    ws.cell(row=row_num, column=target_col).value = correct_val
                    ws.cell(row=row_num, column=target_col).fill  = dvv_fill
                else:
                    new_row = data_start_row + meta["populated_rows"]
                    if issuer_id_col:
                        ws.cell(row=new_row, column=issuer_id_col).value = issuer_id
                    ws.cell(row=new_row, column=serial_id_col).value = dvv_series
                    ws.cell(row=new_row, column=target_col).value = correct_val
                    ws.cell(row=new_row, column=target_col).fill  = dvv_fill
                    row_map[dvv_series] = new_row
                    meta["populated_rows"] += 1
                    logger.info(
                        f"  DVV new row: sheet='{sheet_name}' "
                        f"SeriesID={dvv_series} DataLib={datalib}"
                    )
            else:
                continue

            sheet_overrides += 1

        if sheet_overrides:
            logger.info(
                f"  DVV '{template_key}' / '{sheet_name}': "
                f"{sheet_overrides} override(s) applied"
            )
        total_overrides += sheet_overrides

    logger.info(f"DVV overrides '{template_key}': {total_overrides} total override(s)")
    return total_overrides


# ==============================================================================
# SECTION 6: OUTPUT FORMATTING  (unchanged from v13)
# ==============================================================================

def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font() -> Font:
    return Font(name=CONFIG["HEADER_FONT_NAME"], size=CONFIG["HEADER_FONT_SIZE"],
                bold=True, color=CONFIG["HEADER_FONT_COLOR"])

def _hdr_fill() -> PatternFill:
    return PatternFill("solid", start_color=CONFIG["HEADER_FILL_COLOR"],
                       end_color=CONFIG["HEADER_FILL_COLOR"])

def _dat_font() -> Font:
    return Font(name=CONFIG["DATA_FONT_NAME"], size=CONFIG["DATA_FONT_SIZE"], bold=False)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def format_worksheet(ws, data_start_row: int, populated_rows: int,
                     logger: logging.Logger) -> None:
    """Header styling, data borders, AutoFilter, freeze panes, auto-fit widths."""
    max_col = ws.max_column
    if not max_col:
        return

    last_row = data_start_row + populated_rows - 1
    if populated_rows == 0:
        last_row = data_start_row - 1

    border = _thin_border()

    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        if cell.value is not None:
            cell.font      = _hdr_font()
            cell.fill      = _hdr_fill()
            cell.alignment = _center()
            cell.border    = border

    for r in range(data_start_row, last_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font   = _dat_font()
            cell.border = border  # preserve DVV green fill

    ws.freeze_panes = "A2"

    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(cell.value)) for cell in col_cells if cell.value),
                      default=8)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # AutoFilter instead of Excel Table objects (avoids file-corruption repair).
    filter_ref = f"A1:{get_column_letter(max_col)}1"
    if populated_rows > 0:
        filter_ref = f"A1:{get_column_letter(max_col)}{last_row}"
    ws.auto_filter.ref = filter_ref


# ==============================================================================
# SECTION 7: VALIDATION  (unchanged from v13)
# ==============================================================================

def build_validation_log(all_template_headers: dict, ext_df: pd.DataFrame,
                         issuer_id: str, logger: logging.Logger) -> list:
    """Find template DataLibs missing in extraction and extraction DataLibs unused."""
    ext_cols      = set(ext_df.columns.tolist())
    all_tmpl_cols: set = set()
    records = []

    for tmpl_key, sheets in all_template_headers.items():
        for sheet_name, headers in sheets.items():
            for h in headers:
                if h is None:
                    continue
                all_tmpl_cols.add(h)
                if (h not in ext_cols
                        and h != CONFIG["ISSUER_ID_FIELD"]
                        and h != "serial_id"):
                    records.append({
                        "Check Type": "Template DataLib missing in Extraction",
                        "Template":   tmpl_key,
                        "Sheet":      sheet_name,
                        "Data Lib":   h,
                        "Detail":     "Header in template; no matching column in CSV",
                    })

    for col in ext_cols:
        if col not in all_tmpl_cols and col not in {CONFIG["ISSUER_ID_FIELD"], "serial_id"}:
            records.append({
                "Check Type": "Extraction DataLib unused in Templates",
                "Template":   "ALL",
                "Sheet":      "ALL",
                "Data Lib":   col,
                "Detail":     "Column in CSV; not used in any template sheet",
            })

    logger.info(f"Validation: {len(records)} issues found")
    return records


# ==============================================================================
# SECTION 8: WEB LOGGING HELPER
# Captures log records into a list so they can be shown on the page.
# ==============================================================================

class _ListLogHandler(logging.Handler):
    def __init__(self):
        super().__init__()
        self.lines: list[str] = []

    def emit(self, record):
        try:
            self.lines.append(self.format(record))
        except Exception:
            pass


def build_capture_logger():
    """Return (logger, handler). Read handler.lines after the run for display."""
    logger = logging.getLogger("AnnualUpdateWeb")
    logger.handlers.clear()
    logger.setLevel(logging.DEBUG)
    h = _ListLogHandler()
    h.setLevel(logging.INFO)
    h.setFormatter(logging.Formatter("%(levelname)-8s | %(message)s"))
    logger.addHandler(h)
    return logger, h


# ==============================================================================
# SECTION 9: PIPELINE (web edition)
# Reuses all v13 functions. Inputs come from uploads; outputs are in-memory.
# ==============================================================================

def detect_issuer_id(filename: str) -> str:
    """Pull an IID<digits> token out of a filename, if present."""
    m = re.search(r"(IID\d+)", filename or "", re.IGNORECASE)
    return m.group(1) if m else ""


def run_pipeline(issuer_id: str,
                 extraction_file: tuple[str, bytes],
                 dvv_file: tuple[str, bytes],
                 template_files: dict,
                 logger: logging.Logger):
    """
    Execute the full pipeline.

    Parameters
    ----------
    issuer_id        : e.g. 'IID000000002826074'
    extraction_file  : (filename, bytes) of the extraction CSV
    dvv_file         : (filename, bytes) of the DVV Merged XLSX
    template_files   : { tmpl_key: (filename, bytes) } for each template

    Returns
    -------
    summary    : dict
    outputs    : { output_filename: bytes }
    val_records: list[dict]
    """
    start_time = datetime.datetime.now()
    summary = {
        "Start Time":        start_time.strftime("%Y-%m-%d %H:%M:%S"),
        "End Time":          "",
        "Issuer ID":         issuer_id,
        "Issuer Name":       "",
        "Extraction File":   extraction_file[0],
        "DVV File":          dvv_file[0],
        "Records Processed": 0,
        "Fields Populated":  0,
        "DVV Overrides":     0,
        "Validation Issues": 0,
        "Errors":            [],
        "Status":            "PENDING",
    }
    outputs: dict = {}
    val_records: list = []

    # Issuer name from the CSV filename: "IID..._Aurubis AG.csv" -> "Aurubis AG"
    parts = Path(extraction_file[0]).stem.split("_", 1)
    issuer_name = parts[1].strip() if len(parts) == 2 else issuer_id
    summary["Issuer Name"] = issuer_name

    tmp = Path(tempfile.mkdtemp(prefix="annual_update_"))
    try:
        logger.info("=" * 70)
        logger.info(f"Issuer ID   : {issuer_id}")
        logger.info(f"Issuer Name : {issuer_name}")

        # ── Save inputs to temp disk (functions expect file paths) ───────────
        ext_path = tmp / Path(extraction_file[0]).name
        ext_path.write_bytes(extraction_file[1])
        dvv_path = tmp / Path(dvv_file[0]).name
        dvv_path.write_bytes(dvv_file[1])

        # ── Load extraction CSV ──────────────────────────────────────────────
        logger.info("=" * 70)
        logger.info("STEP 1: Loading extraction CSV")
        ext_df = load_extraction(ext_path, issuer_id, logger)
        summary["Records Processed"] = len(ext_df)

        # ── Load DVV file ────────────────────────────────────────────────────
        logger.info("=" * 70)
        logger.info("STEP 2: Loading DVV file")
        dvv_df = load_dvv(dvv_path, logger)

        # ── Populate templates + DVV overrides + format + save ──────────────
        logger.info("=" * 70)
        logger.info("STEP 3: Populating templates and applying DVV overrides")

        all_template_headers: dict = {}
        total_overrides = 0
        total_fields    = 0

        for tmpl_key, (tname, tbytes) in template_files.items():
            logger.info(f"  ── {tmpl_key}: {tname} ──")
            tpath = tmp / Path(tname).name
            tpath.write_bytes(tbytes)

            try:
                wb, sheet_meta = populate_template(
                    tpath, tmpl_key, ext_df, issuer_id, logger
                )
            except Exception as e:
                msg = f"Population failed for '{tmpl_key}': {e}"
                logger.error(msg)
                logger.debug(traceback.format_exc())
                summary["Errors"].append(msg)
                continue

            all_template_headers[tmpl_key] = {
                sn: m["headers"] for sn, m in sheet_meta.items()
            }
            for m in sheet_meta.values():
                total_fields += len(m["headers"]) * m["populated_rows"]

            try:
                total_overrides += apply_dvv_overrides(
                    wb, sheet_meta, dvv_df, issuer_id, tmpl_key, logger
                )
            except Exception as e:
                msg = f"DVV override failed for '{tmpl_key}': {e}"
                logger.error(msg)
                logger.debug(traceback.format_exc())
                summary["Errors"].append(msg)

            try:
                for sn in wb.sheetnames:
                    if sn in sheet_meta:
                        m = sheet_meta[sn]
                        format_worksheet(
                            wb[sn], m["data_start_row"], m["populated_rows"], logger
                        )
            except Exception as e:
                logger.error(f"Formatting failed for '{tmpl_key}': {e}")
                logger.debug(traceback.format_exc())

            # Save workbook to memory
            try:
                buf = io.BytesIO()
                wb.save(buf)
                out_name = f"{issuer_id}_{CONFIG['OUTPUT_NAMES'][tmpl_key]}.xlsx"
                outputs[out_name] = buf.getvalue()
                logger.info(f"  Generated: {out_name}")
            except Exception as e:
                msg = f"Save failed for '{tmpl_key}': {e}"
                logger.error(msg)
                summary["Errors"].append(msg)

        # ── Validation (logged + returned for display) ──────────────────────
        logger.info("=" * 70)
        logger.info("STEP 4: Validation check")
        val_records = build_validation_log(
            all_template_headers, ext_df, issuer_id, logger
        )
        summary["Validation Issues"] = len(val_records)
        summary["DVV Overrides"]     = total_overrides
        summary["Fields Populated"]  = total_fields

        summary["Status"] = "FAILED" if summary["Errors"] else "SUCCESS"

    except Exception as e:
        logger.error(f"FATAL: {e}")
        logger.debug(traceback.format_exc())
        summary["Errors"].append(str(e))
        summary["Status"] = "FAILED"
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
        end_time = datetime.datetime.now()
        summary["End Time"] = end_time.strftime("%Y-%m-%d %H:%M:%S")
        summary["_elapsed"] = (end_time - start_time).total_seconds()

    return summary, outputs, val_records


# ==============================================================================
# SECTION 10: STREAMLIT UI
# ==============================================================================

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _load_bundled_templates():
    """Return {tmpl_key: (filename, bytes)} for any templates found on disk."""
    found = {}
    if TEMPLATES_DIR.exists():
        for key, fname in CONFIG["TEMPLATES"].items():
            fpath = TEMPLATES_DIR / fname
            if fpath.exists():
                found[key] = (fname, fpath.read_bytes())
    return found


def main():
    st.set_page_config(
        page_title="MSCI Annual Update Processor",
        page_icon="📊",
        layout="wide",
    )

    # Light MSCI-style accent
    st.markdown(
        """
        <style>
          .au-title { color:#0070C0; font-weight:700; margin-bottom:0; }
          .au-sub   { color:#5f6b7a; margin-top:2px; }
          div.stButton > button[kind="primary"] { background:#0070C0; border:0; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<h1 class="au-title">📊 Annual Update Factual Process</h1>',
                unsafe_allow_html=True)
    st.markdown('<p class="au-sub">Upload the extraction CSV and DVV merged file, '
                'then generate the populated bulk-upload templates.</p>',
                unsafe_allow_html=True)
    st.divider()

    # ── Session defaults ─────────────────────────────────────────────────────
    if "issuer_id" not in st.session_state:
        st.session_state.issuer_id = ""
    if "results" not in st.session_state:
        st.session_state.results = None

    # ── INPUTS ───────────────────────────────────────────────────────────────
    left, right = st.columns(2, gap="large")

    with left:
        st.subheader("1 · Inputs")

        ext_upload = st.file_uploader(
            "Extraction CSV", type=["csv"], key="ext_upload",
            help="The raw extraction file, e.g. IID000000002826074_Aurubis AG.csv",
        )
        dvv_upload = st.file_uploader(
            "DVV Merged XLSX", type=["xlsx"], key="dvv_upload",
            help="The DVV merged workbook, e.g. DVV (CG-AU) - 2026-04-17_IID..._Merged.xlsx",
        )

        # Auto-detect Issuer ID from the CSV filename
        detected = detect_issuer_id(ext_upload.name) if ext_upload else ""

        def _use_detected():
            st.session_state.issuer_id = detected

        st.text_input(
            "Issuer ID *", key="issuer_id",
            placeholder="IID000000002826074",
            help="Auto-filled from the CSV filename when possible — edit if needed.",
        )
        if detected and detected != st.session_state.issuer_id:
            st.caption(f"Detected in filename: **{detected}**")
            st.button(f"Use {detected}", on_click=_use_detected)

    with right:
        st.subheader("2 · Templates")

        bundled = _load_bundled_templates()
        template_files: dict = {}

        if len(bundled) == len(CONFIG["TEMPLATES"]):
            st.success(f"Using bundled templates from `{TEMPLATES_DIR}`")
            for key, (fname, _) in bundled.items():
                st.caption(f"• {key} → {fname}")
            template_files = bundled
        else:
            if bundled:
                st.info(
                    f"Found {len(bundled)}/{len(CONFIG['TEMPLATES'])} bundled "
                    "templates. Upload the missing ones below."
                )
            else:
                st.caption(
                    "No bundled templates found. Upload the 3 template files "
                    "(or place them in a `templates/` folder next to this app)."
                )

            labels = {
                "Position":       "Position template (Position_Tab_Split.xlsx)",
                "Scalar_Series1": "Scalar & Series 1 template (Scalar & Series 1_DP.xlsx)",
                "Series2":        "Series 2 template (Series 2- Delete blank cell_DP.xlsx)",
            }
            for key, label in labels.items():
                if key in bundled:
                    template_files[key] = bundled[key]
                    st.caption(f"✓ {key} (bundled)")
                    continue
                up = st.file_uploader(label, type=["xlsx"], key=f"tmpl_{key}")
                if up is not None:
                    template_files[key] = (up.name, up.getvalue())

    st.divider()

    # ── RUN ──────────────────────────────────────────────────────────────────
    issuer_id = st.session_state.issuer_id.strip()
    ready = bool(
        ext_upload and dvv_upload and issuer_id
        and len(template_files) == len(CONFIG["TEMPLATES"])
    )

    if not ready:
        missing = []
        if not ext_upload:
            missing.append("Extraction CSV")
        if not dvv_upload:
            missing.append("DVV file")
        if not issuer_id:
            missing.append("Issuer ID")
        if len(template_files) != len(CONFIG["TEMPLATES"]):
            missing.append(
                f"{len(CONFIG['TEMPLATES']) - len(template_files)} template(s)"
            )
        st.info("Provide the following to enable processing: " + ", ".join(missing))

    if issuer_id and not re.match(r"^IID\d+$", issuer_id, re.IGNORECASE):
        st.warning(
            f"'{issuer_id}' doesn't match the expected pattern 'IID' + digits. "
            "You can still proceed."
        )

    run_clicked = st.button(
        "▶ Run processing", type="primary", disabled=not ready,
        use_container_width=True,
    )

    if run_clicked and ready:
        logger, handler = build_capture_logger()
        with st.spinner("Processing… populating templates and applying DVV overrides"):
            summary, outputs, val_records = run_pipeline(
                issuer_id=issuer_id,
                extraction_file=(ext_upload.name, ext_upload.getvalue()),
                dvv_file=(dvv_upload.name, dvv_upload.getvalue()),
                template_files=template_files,
                logger=logger,
            )
        st.session_state.results = {
            "summary":     summary,
            "outputs":     outputs,
            "val_records": val_records,
            "log":         handler.lines,
        }

    # ── RESULTS (rendered from session_state so downloads don't re-run) ──────
    results = st.session_state.results
    if results:
        summary     = results["summary"]
        outputs     = results["outputs"]
        val_records = results["val_records"]
        log_lines   = results["log"]

        st.divider()
        st.subheader("3 · Results")

        if summary["Status"] == "SUCCESS":
            st.success(f"Completed for **{summary['Issuer Name']}** "
                       f"in {summary.get('_elapsed', 0):.1f}s")
        else:
            st.error("Completed with errors — see details below.")

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Records processed", summary["Records Processed"])
        m2.metric("Fields populated",  summary["Fields Populated"])
        m3.metric("DVV overrides",     summary["DVV Overrides"])
        m4.metric("Validation issues", summary["Validation Issues"])

        if summary["Errors"]:
            with st.expander(f"⚠ Errors ({len(summary['Errors'])})", expanded=True):
                for err in summary["Errors"]:
                    st.write(f"• {err}")

        # Downloads
        if outputs:
            st.markdown("**Download outputs**")
            cols = st.columns(len(outputs))
            for col, (fname, data) in zip(cols, outputs.items()):
                with col:
                    st.download_button(
                        f"⬇️ {fname}", data=data, file_name=fname,
                        mime=XLSX_MIME, use_container_width=True,
                    )

            # ZIP of everything
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for fname, data in outputs.items():
                    zf.writestr(fname, data)
            st.download_button(
                "⬇️ Download all (ZIP)", data=zip_buf.getvalue(),
                file_name=f"{summary['Issuer ID']}_outputs.zip",
                mime="application/zip", type="primary",
                use_container_width=True,
            )

        # Validation table
        if val_records:
            with st.expander(f"Validation issues ({len(val_records)})"):
                st.dataframe(pd.DataFrame(val_records), use_container_width=True,
                             hide_index=True)
        else:
            st.caption("No validation issues found.")

        # Execution log
        with st.expander("Execution log"):
            st.code("\n".join(log_lines) or "(no log output)", language="text")


if __name__ == "__main__":
    main()
