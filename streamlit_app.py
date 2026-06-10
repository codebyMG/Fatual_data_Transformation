"""
================================================================================
MSCI Annual Update Factual Process Automation  —  Streamlit Web App
================================================================================
Version     : 13.0.0 (Streamlit)
Based on    : annual_update_processor v13.0.0 (command-line)

Changes in 13.0.0 (integrates the v13 command-line logic)
  - SECTION-AWARE Series ID resolution. The extraction CSV contains MULTIPLE
    "Series ID" columns (one per section: directors, committees, ownership …).
    pandas loads these as 'Series ID', 'Series ID.1', 'Series ID.2', …  Each
    data-lib column is now matched to the NEAREST "Series ID" column that sits
    BEFORE it (build_datalib_to_series_map). This replaces the old single
    fixed-column approach (Column E / index 4), which was wrong: in the current
    files Column E is PRIMARY_DISCLOSURE_LANG, not a Series ID.
  - Series-sheet population now uses two gates:
        Gate 1 — the section's Series ID for the row must be non-blank;
        Gate 2 — at least one DATA column belonging to that section must be
                 non-blank (stops rows that carry only a Series ID, e.g.
                 directors with no committee data, from being written).
  - Scalar DVV override writes directly to the single issuer row (row 2),
    matched by DATALIB_TAG only — no Series ID matching.
  - Series-sheet DVV override matches on serial_id + DATALIB_TAG and, when no
    matching row exists, CREATES A NEW ROW (issuer id + serial id + value,
    green-highlighted) instead of logging an exception.

Retained from the earlier Streamlit build
  - DVV columns are matched by HEADER NAME (DATALIB_TAG / CORRECT_VALUE /
    SERIALID / TAB), not fixed Excel column letters, so a 'UUID4' column
    inserted at position A (which shifted every column right by one) does not
    break matching. Matching is case-insensitive and whitespace-tolerant.
  - DVV rows are filtered in two stages: keep non-blank CORRECT_VALUE, then
    drop rows whose TAB is Individual / Director Data - Board /
    Director Attributes / Positions.

What this app does
  Upload the extraction CSV and the DVV merged file, press Run, and download
  the populated bulk-upload templates as a ZIP.

Key behaviour (preserved)
  - THREE fixed templates. Scalar and Series 1 are combined into one file.
  - The scalar (issuer-level) sheet is detected by SHEET NAME ("Scalar"),
    not by file. Every other sheet is series-level (one row per CSV row).
  - No openpyxl Table objects. AutoFilter is used instead (Excel Tables made
    openpyxl emit orphaned table XML, triggering the "We found a problem"
    repair dialog). AutoFilter gives the same column dropdown / sort UX.
  - The output contains ONLY the three populated template files.

Web-app specifics
  - Templates are FIXED: bundled with the app (committed to the repo's
    'templates/' folder) and loaded automatically. Only CSV + DVV are uploaded.
  - Issuer ID is read automatically from the DMX_ISSUER_ID column of the CSV.
    One issuer  -> processed on its own.  Several issuers -> each processed and
    all outputs bundled together (per-issuer subfolders inside the ZIP).
  - Everything is handled in memory (BytesIO); nothing is written to disk.
  - Validation issues and the DVV audit trail are shown on-screen and offered
    as optional CSV downloads (they are intentionally NOT inside the ZIP, to
    match the "three files only" output).
Requirements: streamlit | pandas | openpyxl
================================================================================
"""

import io
import re
import zipfile
import logging
import datetime
import traceback
from pathlib import Path

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# Table / TableStyleInfo intentionally NOT imported — see module docstring.


# ==============================================================================
# SECTION 1: CONFIGURATION
# Templates are FIXED: bundled with the app (committed to the repo) and loaded
# automatically. Only the CSV and DVV are uploaded per run.
# ==============================================================================

CONFIG = {
    # Folder (relative to this script) holding the fixed template files.
    "TEMPLATE_DIR": "templates",

    # Fixed templates: internal key -> filename inside TEMPLATE_DIR.
    # Commit these three files to the repo's 'templates/' folder with these
    # exact names. Scalar and Series 1 are combined in one file.
    "TEMPLATE_FILES": {
        "Position":       "Position_Tab_Split.xlsx",
        "Scalar_Series1": "Scalar_Series1_DP.xlsx",
        "Series2":        "Series2_DP.xlsx",
    },

    # Display labels for the status panel.
    "TEMPLATE_LABELS": {
        "Position":       "Position",
        "Scalar_Series1": "Scalar & Series 1",
        "Series2":        "Series 2",
    },

    # Output file-name suffixes per template key.
    # Output files are named:  IssuerID_<suffix>.xlsx
    # These keep the names the downstream bulk-upload process expects.
    "OUTPUT_NAMES": {
        "Position":       "Position_Tab_Split",
        "Scalar_Series1": "Scalar & Series 1_DP",
        "Series2":        "Series 2- Delete blank cell_DP",
    },

    # Name of the scalar (issuer-level) sheet inside the combined file.
    # All other sheets in any template are treated as series-level.
    "SCALAR_SHEET_NAME": "Scalar",

    # DVV columns matched by HEADER NAME (robust to column insertion/removal
    # upstream). Earlier versions matched by fixed Excel column letters
    # (F / AG / AL); a new 'UUID4' column inserted at position A shifted every
    # column by one and broke that approach. Header-name matching survives such
    # changes as long as these header names exist in the DVV file's row 1.
    "DVV_DATALIB_HEADER":    "DATALIB_TAG",    # -> Data Lib
    "DVV_CORRECTVAL_HEADER": "CORRECT_VALUE",  # -> Correct Value
    "DVV_SERIES_HEADER":     "SERIALID",       # -> Series ID
    "DVV_TAB_HEADER":        "TAB",            # -> TAB (used for filtering)

    # TAB values to EXCLUDE from DVV overrides. Rows whose TAB is any of these
    # are dropped after the CORRECT_VALUE filter and never applied to the
    # templates. (Individual / director-attribute / position rows are handled
    # elsewhere and must not override the bulk-upload templates.)
    "DVV_EXCLUDED_TABS": {
        "Individual",
        "Director Data - Board",
        "Director Attributes",
        "Positions",
    },

    # ── Extraction Series ID ──────────────────────────────────────────────────
    # The extraction CSV has MULTIPLE "Series ID" columns (one per section).
    # Series IDs are resolved per data-lib via build_datalib_to_series_map(),
    # NOT from a single fixed column. The index below is only a fallback /
    # diagnostic hint used when logging; it is no longer the source of truth.
    # (In the current files Column E / index 4 is PRIMARY_DISCLOSURE_LANG.)
    "EXTRACTION_SERIES_COL_INDEX": 29,

    # Field names.
    "ISSUER_ID_FIELD": "DMX_ISSUER_ID",

    # Formatting.
    "HEADER_FONT_NAME":    "Times New Roman",
    "HEADER_FONT_SIZE":    12,
    "HEADER_FONT_COLOR":   "FFFFFF",   # White text
    "HEADER_FILL_COLOR":   "0070C0",   # Blue fill
    "DATA_FONT_NAME":      "Times New Roman",
    "DATA_FONT_SIZE":      12,
    "DVV_HIGHLIGHT_COLOR": "E6FDCF",   # Light green for DVV-updated cells
}


# ==============================================================================
# SECTION 2: LOGGING (in-memory)
# ==============================================================================

def setup_logging() -> tuple[logging.Logger, io.StringIO]:
    """Log to an in-memory buffer so the messages can be shown in the app."""
    log_buffer = io.StringIO()

    logger = logging.getLogger("AnnualUpdate")
    logger.handlers.clear()
    logger.setLevel(logging.DEBUG)

    handler = logging.StreamHandler(log_buffer)
    handler.setLevel(logging.INFO)
    handler.setFormatter(logging.Formatter("%(levelname)-8s | %(message)s"))
    logger.addHandler(handler)
    logger.propagate = False
    return logger, log_buffer


def _candidate_dirs() -> list[Path]:
    """
    Directories searched for the fixed template files, in priority order:
    the app's own folder (repo root) first, then a 'templates/' subfolder.
    This works whether the files are committed at the repo root or inside
    a 'templates/' folder.
    """
    base = Path(__file__).resolve().parent
    return [base, base / CONFIG["TEMPLATE_DIR"]]


def load_fixed_templates(logger: logging.Logger) -> tuple[dict[str, bytes], list[str]]:
    """
    Load the fixed template files bundled with the app.
    Returns (templates_bytes_by_key, list_of_missing_keys).
    """
    candidates = _candidate_dirs()
    templates: dict[str, bytes] = {}
    missing: list[str] = []

    for key, filename in CONFIG["TEMPLATE_FILES"].items():
        found_path = next(
            (d / filename for d in candidates if (d / filename).exists()), None
        )
        if found_path is not None:
            templates[key] = found_path.read_bytes()
            logger.info(f"Loaded fixed template '{key}': {found_path}")
        else:
            missing.append(key)
            logger.warning(
                f"Fixed template '{key}' ({filename}) not found in: "
                + ", ".join(str(d) for d in candidates)
            )

    return templates, missing


# ==============================================================================
# SECTION 3: EXTRACTION CSV PROCESSING
# ==============================================================================

def standardize_header(raw_header: str) -> str:
    """
    Convert a CSV header to its Data Lib identifier.

    Format 1 - already a Data Lib (no brackets):
        DMX_ISSUER_ID  ->  DMX_ISSUER_ID
    Format 2 - human label with Data Lib in final parentheses:
        Audit Board Member (REL_AUDIT_BOARD)  ->  REL_AUDIT_BOARD

    Rule: extract text inside the LAST pair of parentheses; otherwise unchanged.
    "Series ID" (and pandas' duplicate suffixes 'Series ID.1', …) have no
    parentheses, so they pass through unchanged and remain detectable as the
    section markers used by build_datalib_to_series_map().
    """
    raw = str(raw_header).strip()
    match = re.search(r"\(([^)]+)\)\s*$", raw)
    if match:
        return match.group(1).strip()
    return raw


def load_extraction(file_bytes: bytes, display_name: str,
                    logger: logging.Logger) -> pd.DataFrame:
    """
    Load extraction CSV (from uploaded bytes) and standardise headers.
    All values read as strings to prevent type inference. No column reordering.

    The raw CSV carries several "Series ID" columns (one per section). pandas
    auto-suffixes duplicates as 'Series ID.1', 'Series ID.2', … — these are
    preserved here and consumed later by build_datalib_to_series_map().
    """
    logger.info(f"Loading extraction CSV: {display_name}")
    try:
        df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="utf-8-sig")
    except UnicodeDecodeError:
        df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, encoding="latin-1")

    if df.empty:
        raise ValueError(f"Extraction CSV is empty: {display_name}")

    df.columns = [standardize_header(c) for c in df.columns]

    # Report the Series ID columns that were found (diagnostic only).
    series_cols = [
        c for c in df.columns
        if c == "Series ID" or re.match(r"^Series ID(\.\d+)?$", c)
    ]
    if series_cols:
        logger.info(
            f"Found {len(series_cols)} Series ID column(s): "
            f"{', '.join(series_cols)}"
        )
    else:
        logger.warning(
            "No 'Series ID' column found in the extraction CSV. "
            "Series sheets will be left empty (no Series ID to match)."
        )

    logger.info(
        f"Extraction loaded: {len(df)} rows, {len(df.columns)} columns "
        "after header standardisation"
    )
    return df


def build_datalib_to_series_map(ext_df: pd.DataFrame) -> dict[str, int]:
    """
    Build a mapping: datalib_code -> Series ID column index in ext_df.

    The extraction CSV has multiple 'Series ID' columns (one per section).
    Each Series ID column immediately precedes the data-lib columns for that
    section. For any given data-lib code, the correct Series ID is the one
    whose column position is the closest one BEFORE that code's column.

    Returns e.g. { 'DIRGENDER': 7, 'COMMITTEENAME': 10, ... } where values are
    0-based column indices into ext_df.
    """
    cols = list(ext_df.columns)
    series_id_positions = [
        i for i, c in enumerate(cols)
        if c == "Series ID" or re.match(r"^Series ID(\.\d+)?$", c)
    ]

    datalib_to_series_col: dict[str, int] = {}
    for col_idx, col_name in enumerate(cols):
        if col_name == "Series ID" or re.match(r"^Series ID(\.\d+)?$", col_name):
            continue
        before = [p for p in series_id_positions if p < col_idx]
        if before:
            datalib_to_series_col[col_name] = before[-1]

    return datalib_to_series_col


def row_has_data(ext_row: pd.Series, template_headers: list[str]) -> bool:
    """True if the row has at least one non-blank value for a template header
    (excluding DMX_ISSUER_ID and serial_id). Prevents writing empty rows."""
    skip = {CONFIG["ISSUER_ID_FIELD"], "serial_id"}
    for h in template_headers:
        if h in skip:
            continue
        val = ext_row.get(h)
        if pd.notna(val) and str(val).strip() not in ("", "nan"):
            return True
    return False


def detect_issuer_ids(ext_df: pd.DataFrame,
                      logger: logging.Logger) -> list[str]:
    """Return the list of distinct, non-blank issuer IDs in the CSV."""
    col = CONFIG["ISSUER_ID_FIELD"]
    if col not in ext_df.columns:
        raise ValueError(
            f"Column '{col}' not found in the extraction CSV. "
            "The issuer ID cannot be determined automatically."
        )
    series = ext_df[col].astype(str).str.strip()
    ids = [v for v in series.unique().tolist() if v and v.lower() != "nan"]
    if not ids:
        raise ValueError(
            f"No issuer IDs found in column '{col}'. The column is empty."
        )
    logger.info(f"Detected {len(ids)} issuer ID(s): {', '.join(ids)}")
    return ids


# ==============================================================================
# SECTION 4: DVV FILE PROCESSING
# ==============================================================================

def load_dvv(file_bytes: bytes, display_name: str,
             logger: logging.Logger) -> pd.DataFrame:
    """
    Load DVV Merged XLSX (from uploaded bytes) and return the rows to apply.

    Columns are matched by HEADER NAME (row 1), not fixed Excel column letters,
    so the function is robust to columns being inserted/removed upstream
    (e.g. the 'UUID4' column added at position A). Matching is case-insensitive
    and whitespace-tolerant.

    Two filters are applied, in this order:
      1. Keep only rows where CORRECT_VALUE is non-blank.
      2. Drop rows whose TAB is in CONFIG["DVV_EXCLUDED_TABS"]
         (Individual, Director Data - Board, Director Attributes, Positions).

    The surviving rows split into:
      - scalar rows  : SERIALID is blank  -> applied to the Scalar sheet by
                       DATALIB_TAG only (handled in apply_dvv_overrides).
      - series rows  : SERIALID is present -> applied to series sheets by
                       SERIALID + DATALIB_TAG.
    """
    logger.info(f"Loading DVV file: {display_name}")

    try:
        dvv_df = pd.read_excel(
            io.BytesIO(file_bytes), sheet_name=0, dtype=str,
            header=0, engine="openpyxl"
        )
    except Exception as e:
        raise RuntimeError(f"Cannot read DVV file '{display_name}': {e}")

    if dvv_df.empty:
        raise ValueError(f"DVV file is empty: {display_name}")

    # Build a case-insensitive, whitespace-tolerant lookup of actual headers.
    norm_to_actual = {
        str(c).strip().upper(): c for c in dvv_df.columns
    }

    wanted = {
        CONFIG["DVV_DATALIB_HEADER"]:    "_DVV_DATALIB",
        CONFIG["DVV_CORRECTVAL_HEADER"]: "_DVV_CORRECT_VALUE",
        CONFIG["DVV_SERIES_HEADER"]:     "_DVV_SERIES_ID",
        CONFIG["DVV_TAB_HEADER"]:        "_DVV_TAB",
    }

    rename_map: dict = {}
    missing: list[str] = []
    for header_name, internal_name in wanted.items():
        actual = norm_to_actual.get(header_name.strip().upper())
        if actual is None:
            missing.append(header_name)
        else:
            rename_map[actual] = internal_name

    if missing:
        raise ValueError(
            "DVV file is missing expected column header(s): "
            + ", ".join(missing)
            + ". Found headers: "
            + ", ".join(str(c) for c in dvv_df.columns)
        )

    dvv_df = dvv_df.rename(columns=rename_map)
    logger.info(
        "DVV columns matched by header -> "
        f"DataLib='{CONFIG['DVV_DATALIB_HEADER']}' "
        f"CorrectVal='{CONFIG['DVV_CORRECTVAL_HEADER']}' "
        f"SeriesID='{CONFIG['DVV_SERIES_HEADER']}' "
        f"TAB='{CONFIG['DVV_TAB_HEADER']}'"
    )

    # ── Filter 1: keep only rows that HAVE a Correct Value ────────────────────
    total = len(dvv_df)
    dvv_df = dvv_df[
        dvv_df["_DVV_CORRECT_VALUE"].notna()
        & (dvv_df["_DVV_CORRECT_VALUE"].str.strip() != "")
    ].copy()
    after_cv = len(dvv_df)
    logger.info(f"DVV filter 1 (Correct Value present): {total} -> {after_cv} rows")

    # ── Filter 2: drop excluded TAB categories ────────────────────────────────
    # Case-insensitive, whitespace-tolerant comparison against the excluded set.
    excluded_norm = {t.strip().upper() for t in CONFIG["DVV_EXCLUDED_TABS"]}
    tab_norm = dvv_df["_DVV_TAB"].fillna("").astype(str).str.strip().str.upper()
    dvv_df = dvv_df[~tab_norm.isin(excluded_norm)].copy()
    after_tab = len(dvv_df)
    logger.info(
        f"DVV filter 2 (excluded TABs {sorted(CONFIG['DVV_EXCLUDED_TABS'])}): "
        f"{after_cv} -> {after_tab} rows"
    )

    # Split for logging: scalar rows (no Series ID) vs series rows (Series ID).
    sid_norm = dvv_df["_DVV_SERIES_ID"].fillna("").astype(str).str.strip().str.lower()
    n_scalar = int((sid_norm.isin(["", "nan"])).sum())
    n_series = after_tab - n_scalar
    logger.info(
        f"DVV rows to apply: {after_tab} total "
        f"({n_scalar} scalar / no Series ID, {n_series} series / with Series ID)"
    )
    return dvv_df


# ==============================================================================
# SECTION 5: TEMPLATE POPULATION
# ==============================================================================

def populate_template(
    template_bytes: bytes,
    template_key: str,
    ext_df: pd.DataFrame,
    issuer_id: str,
    logger: logging.Logger,
) -> tuple[openpyxl.Workbook, dict]:
    """
    Populate all sheets in a template workbook from the extraction DataFrame.

    Scalar sheet (sheet named CONFIG["SCALAR_SHEET_NAME"]):
        one issuer-level row, no Series ID, no blank-row check.

    Every other sheet (series-level) uses two gates per extraction row:
        Gate 1 — the section's Series ID for the row must be non-blank;
        Gate 2 — at least one DATA column belonging to that section must be
                 non-blank. This stops rows that carry only a Series ID (e.g.
                 a director with no committee data) from being written.

    Each series sheet finds its OWN section: the Series ID column nearest-before
    the sheet's data-lib headers, and the data columns between that Series ID
    column and the next one. Section membership comes from
    build_datalib_to_series_map().

    The scalar sheet is detected by NAME, so it works whether it lives in its
    own file or is combined with series sheets (the Scalar & Series 1 file).

    Matching: exact Data Lib match only. Unmatched headers -> blank + logged.
    """
    logger.info(f"Populating '{template_key}'")
    wb = load_workbook(io.BytesIO(template_bytes))
    sheet_meta: dict = {}

    # Build once: datalib_code -> which Series ID column index to use.
    datalib_to_series_map = build_datalib_to_series_map(ext_df)
    all_std = list(ext_df.columns)
    series_positions = [
        i for i, c in enumerate(all_std)
        if c == "Series ID" or re.match(r"^Series ID(\.\d+)?$", c)
    ]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        headers = [c.value for c in ws[1] if c.value is not None]
        if not headers:
            logger.warning(f"  Sheet '{sheet_name}' has no headers - skipping.")
            continue

        header_col_map = {
            cell.value: col_idx
            for col_idx, cell in enumerate(ws[1], start=1)
            if cell.value is not None
        }

        # Scalar sheet detected by name (works in combined files).
        is_scalar = (sheet_name == CONFIG["SCALAR_SHEET_NAME"])

        data_start_row = 2
        populated_rows = 0

        # -- SCALAR: one issuer row --------------------------------------------
        if is_scalar:
            if CONFIG["ISSUER_ID_FIELD"] in ext_df.columns:
                issuer_rows = ext_df[
                    ext_df[CONFIG["ISSUER_ID_FIELD"]].astype(str).str.strip()
                    == issuer_id
                ]
                if issuer_rows.empty:
                    issuer_rows = ext_df
            else:
                issuer_rows = ext_df

            if issuer_rows.empty:
                logger.warning(
                    f"  No rows for issuer {issuer_id}. "
                    f"Scalar sheet '{sheet_name}' left empty."
                )
                # Still record meta so downstream steps are consistent.
                sheet_meta[sheet_name] = {
                    "headers":        headers,
                    "header_col_map": header_col_map,
                    "data_start_row": data_start_row,
                    "populated_rows": 0,
                }
                continue

            ext_row   = issuer_rows.iloc[0]
            write_row = data_start_row

            for tmpl_header, col_idx in header_col_map.items():
                if tmpl_header == CONFIG["ISSUER_ID_FIELD"]:
                    ws.cell(row=write_row, column=col_idx).value = issuer_id
                elif tmpl_header in ext_df.columns:
                    val = ext_row.get(tmpl_header)
                    ws.cell(row=write_row, column=col_idx).value = (
                        None if (pd.isna(val) or str(val).strip() in ("", "nan"))
                        else str(val).strip()
                    )

            populated_rows = 1

        # -- POSITION / SERIES: one row per extraction row ---------------------
        else:
            # Find the Series ID column for this sheet (nearest-before the
            # sheet's data-lib headers) and the data columns in its section.
            sheet_series_col_idx = None
            for h in headers:
                if h in (CONFIG["ISSUER_ID_FIELD"], "serial_id"):
                    continue
                idx = datalib_to_series_map.get(h)
                if idx is not None and sheet_series_col_idx is None:
                    sheet_series_col_idx = idx

            sheet_section_cols: list[str] = []
            if sheet_series_col_idx is not None:
                next_series = next(
                    (p for p in series_positions if p > sheet_series_col_idx),
                    len(all_std)
                )
                sheet_section_cols = [
                    all_std[i]
                    for i in range(sheet_series_col_idx + 1, next_series)
                    if all_std[i] in headers
                ]

            for _, ext_row in ext_df.iterrows():

                # Gate 1: section Series ID must be non-blank.
                if sheet_series_col_idx is not None:
                    sid_col_name = ext_df.columns[sheet_series_col_idx]
                    sid_val = ext_row.get(sid_col_name)
                    if pd.isna(sid_val) or str(sid_val).strip() in ("", "nan"):
                        continue
                    section_series_id = str(sid_val).strip()
                else:
                    section_series_id = None

                # Gate 2: at least one DATA column in this section must be
                # non-blank (prevents rows that only have a Series ID).
                if sheet_section_cols:
                    section_has_data = any(
                        pd.notna(ext_row.get(c))
                        and str(ext_row.get(c, "")).strip() not in ("", "nan")
                        for c in sheet_section_cols
                    )
                    if not section_has_data:
                        continue
                elif not row_has_data(ext_row, headers):
                    continue

                write_row = data_start_row + populated_rows

                for tmpl_header, col_idx in header_col_map.items():
                    if tmpl_header == CONFIG["ISSUER_ID_FIELD"]:
                        ws.cell(row=write_row, column=col_idx).value = issuer_id
                    elif tmpl_header == "serial_id":
                        ws.cell(row=write_row, column=col_idx).value = section_series_id
                    elif tmpl_header in ext_df.columns:
                        val = ext_row.get(tmpl_header)
                        ws.cell(row=write_row, column=col_idx).value = (
                            None if (pd.isna(val) or str(val).strip() in ("", "nan"))
                            else str(val).strip()
                        )

                populated_rows += 1

        sheet_meta[sheet_name] = {
            "headers":        headers,
            "header_col_map": header_col_map,
            "data_start_row": data_start_row,
            "populated_rows": populated_rows,
        }
        logger.debug(f"  Sheet '{sheet_name}': {populated_rows} row(s) written.")

    return wb, sheet_meta


# ==============================================================================
# SECTION 6: DVV OVERRIDE PROCESSING
# ==============================================================================

def apply_dvv_overrides(
    wb: openpyxl.Workbook,
    sheet_meta: dict,
    dvv_df: pd.DataFrame,
    issuer_id: str,
    template_key: str,
    logger: logging.Logger,
) -> list[dict]:
    """
    Apply DVV Correct Value overrides to the populated workbook.

    Routing is driven by the DVV row's Series ID:
      - Scalar sheet (named CONFIG["SCALAR_SHEET_NAME"]): only DVV rows with a
        BLANK Series ID are applied, matched by DATALIB_TAG. The value is
        written directly to the single issuer row (data_start_row). DVV rows
        that carry a Series ID are series-level corrections and are skipped on
        the scalar sheet.
      - Series sheets: only DVV rows WITH a Series ID are applied, matched on
        serial_id + DATALIB_TAG.
            * Match found     -> overwrite that cell + green highlight.
            * No match found  -> CREATE A NEW ROW (issuer id + serial id +
                                 value, green) and record it. (v13 behaviour;
                                 replaces the old "log an exception" path.)

    (The dvv_df passed in has already been filtered to non-blank Correct Values
    with the excluded TABs removed — see load_dvv.)

    Updated/added cells are highlighted with DVV_HIGHLIGHT_COLOR.
    Returns audit records (one per applied change or created row).
    """
    dvv_fill = PatternFill(
        "solid",
        start_color=CONFIG["DVV_HIGHLIGHT_COLOR"],
        end_color=CONFIG["DVV_HIGHLIGHT_COLOR"],
    )
    audit_recs: list[dict] = []
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for sheet_name in wb.sheetnames:
        if sheet_name not in sheet_meta:
            continue

        ws             = wb[sheet_name]
        meta           = sheet_meta[sheet_name]
        header_col_map = meta["header_col_map"]
        data_start_row = meta["data_start_row"]
        is_scalar      = (sheet_name == CONFIG["SCALAR_SHEET_NAME"])

        serial_id_col = header_col_map.get("serial_id")
        issuer_id_col = header_col_map.get(CONFIG["ISSUER_ID_FIELD"])

        # Build serial_id -> excel row number map from existing rows
        # (series sheets only; scalar has no Series ID matching).
        row_map: dict[str, int] = {}
        if not is_scalar and serial_id_col:
            for r in range(data_start_row,
                           data_start_row + meta["populated_rows"]):
                sid = ws.cell(row=r, column=serial_id_col).value
                if sid is not None and str(sid).strip() not in ("", "nan"):
                    row_map[str(sid).strip()] = r

        for _, dvv_row in dvv_df.iterrows():
            datalib     = str(dvv_row.get("_DVV_DATALIB", "")).strip()
            correct_val = dvv_row.get("_DVV_CORRECT_VALUE")

            # Normalise the DVV Series ID; treat NaN / "nan" / "" all as blank.
            raw_series = dvv_row.get("_DVV_SERIES_ID")
            dvv_series = (
                "" if (pd.isna(raw_series)
                       or str(raw_series).strip().lower() in ("", "nan"))
                else str(raw_series).strip()
            )

            # Skip if DATALIB_TAG is not a header in this sheet.
            if not datalib or datalib not in header_col_map:
                continue

            target_col = header_col_map[datalib]

            # ── SCALAR: blank-Series rows only, write to the issuer row ───────
            if is_scalar:
                if dvv_series:
                    # Series-level correction — does not belong on scalar sheet.
                    continue
                cell = ws.cell(row=data_start_row, column=target_col)
                old_value = cell.value
                cell.value = correct_val
                cell.fill  = dvv_fill
                audit_recs.append({
                    "Issuer ID": issuer_id, "Series ID": "",
                    "Data Lib": datalib,
                    "Old Value": str(old_value) if old_value is not None else "",
                    "New Value": str(correct_val), "Update Timestamp": ts,
                    "Template Name": template_key, "Worksheet Name": sheet_name,
                    "Status": "SUCCESS",
                })
                continue

            # ── SERIES: needs a Series ID and a serial_id column ──────────────
            if not dvv_series or not serial_id_col:
                # No Series ID (and not scalar), or sheet has no serial_id
                # column to match against — skip.
                continue

            if dvv_series in row_map:
                # Match found -> overwrite in place.
                row_num   = row_map[dvv_series]
                cell      = ws.cell(row=row_num, column=target_col)
                old_value = cell.value
                cell.value = correct_val
                cell.fill  = dvv_fill
                audit_recs.append({
                    "Issuer ID": issuer_id, "Series ID": dvv_series,
                    "Data Lib": datalib,
                    "Old Value": str(old_value) if old_value is not None else "",
                    "New Value": str(correct_val), "Update Timestamp": ts,
                    "Template Name": template_key, "Worksheet Name": sheet_name,
                    "Status": "SUCCESS",
                })
            else:
                # No matching row -> CREATE A NEW ROW (v13 behaviour).
                new_row = data_start_row + meta["populated_rows"]
                if issuer_id_col:
                    ws.cell(row=new_row, column=issuer_id_col).value = issuer_id
                ws.cell(row=new_row, column=serial_id_col).value = dvv_series
                cell = ws.cell(row=new_row, column=target_col)
                cell.value = correct_val
                cell.fill  = dvv_fill
                row_map[dvv_series] = new_row
                meta["populated_rows"] += 1
                logger.info(
                    f"  DVV new row: sheet='{sheet_name}' "
                    f"SeriesID={dvv_series} DataLib={datalib}"
                )
                audit_recs.append({
                    "Issuer ID": issuer_id, "Series ID": dvv_series,
                    "Data Lib": datalib, "Old Value": "",
                    "New Value": str(correct_val), "Update Timestamp": ts,
                    "Template Name": template_key, "Worksheet Name": sheet_name,
                    "Status": "SUCCESS (new row created)",
                })

    success = sum(1 for r in audit_recs if r["Status"].startswith("SUCCESS"))
    new_rows = sum(1 for r in audit_recs if "new row" in r["Status"])
    logger.info(
        f"DVV overrides '{template_key}': {success} change(s) "
        f"({new_rows} via new rows)"
    )
    return audit_recs


# ==============================================================================
# SECTION 7: OUTPUT FORMATTING  (AutoFilter only — no Table objects)
# ==============================================================================

def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font() -> Font:
    return Font(
        name=CONFIG["HEADER_FONT_NAME"], size=CONFIG["HEADER_FONT_SIZE"],
        bold=True, color=CONFIG["HEADER_FONT_COLOR"]
    )

def _hdr_fill() -> PatternFill:
    return PatternFill(
        "solid",
        start_color=CONFIG["HEADER_FILL_COLOR"],
        end_color=CONFIG["HEADER_FILL_COLOR"]
    )

def _dat_font() -> Font:
    return Font(name=CONFIG["DATA_FONT_NAME"], size=CONFIG["DATA_FONT_SIZE"], bold=False)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def format_worksheet(
    ws,
    data_start_row: int,
    populated_rows: int,
    logger: logging.Logger,
) -> None:
    """
    Header styling, data borders, AutoFilter, freeze top row, auto-fit columns.
    DVV green highlights are preserved (the data loop sets font/border only).

    NOTE: populated_rows is read from sheet_meta AFTER apply_dvv_overrides, so
    any rows the DVV step added are included in the formatted range.

    AutoFilter is used instead of Excel Table objects: openpyxl emits orphaned
    table XML that Excel must repair on open ("We found a problem"). AutoFilter
    gives the same column dropdown / sort UX with zero corruption risk.
    """
    max_col = ws.max_column
    if not max_col:
        return

    last_row = data_start_row + populated_rows - 1
    if populated_rows == 0:
        last_row = data_start_row - 1

    border = _thin_border()

    # Header row
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        if cell.value is not None:
            cell.font      = _hdr_font()
            cell.fill      = _hdr_fill()
            cell.alignment = _center()
            cell.border    = border

    # Data rows (preserve DVV highlight: set font + border only)
    for r in range(data_start_row, last_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font   = _dat_font()
            cell.border = border

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-fit column widths
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value), default=8
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # AutoFilter over the full data range (replaces Excel Table objects).
    filter_ref = f"A1:{get_column_letter(max_col)}1"
    if populated_rows > 0:
        filter_ref = f"A1:{get_column_letter(max_col)}{last_row}"
    ws.auto_filter.ref = filter_ref


# ==============================================================================
# SECTION 8: VALIDATION (in memory — surfaced in the UI, not written to files)
# ==============================================================================

def build_validation_log(
    all_template_headers: dict,
    ext_df: pd.DataFrame,
    issuer_id: str,
    logger: logging.Logger,
) -> list[dict]:
    """Validation records: template Data Libs missing in extraction, and
    extraction Data Libs unused in any template.

    'Series ID' marker columns (and their pandas suffixes) are not real data
    libs, so they are excluded from the 'unused extraction column' check."""
    ext_cols      = set(ext_df.columns.tolist())
    all_tmpl_cols: set[str] = set()
    records = []

    def is_series_marker(name: str) -> bool:
        return name == "Series ID" or bool(re.match(r"^Series ID(\.\d+)?$", name))

    for tmpl_key, sheets in all_template_headers.items():
        for sheet_name, headers in sheets.items():
            for h in headers:
                if h is None:
                    continue
                all_tmpl_cols.add(h)
                if (h not in ext_cols
                        and h != CONFIG["ISSUER_ID_FIELD"]
                        and h != "serial_id"):
                    records.append({
                        "Check Type": "Template DataLib missing in Extraction",
                        "Template": tmpl_key, "Sheet": sheet_name, "Data Lib": h,
                        "Detail": "Header in template; no matching column in CSV",
                    })

    for col in ext_cols:
        if (col not in all_tmpl_cols
                and col not in {CONFIG["ISSUER_ID_FIELD"], "serial_id"}
                and not is_series_marker(col)):
            records.append({
                "Check Type": "Extraction DataLib unused in Templates",
                "Template": "ALL", "Sheet": "ALL", "Data Lib": col,
                "Detail": "Column in CSV; not used in any template sheet",
            })

    logger.info(f"Validation: {len(records)} issues found")
    return records


# ==============================================================================
# SECTION 9: ORCHESTRATION (in memory)
# ==============================================================================

def _wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def process_one_issuer(
    issuer_id: str,
    issuer_ext_df: pd.DataFrame,
    dvv_df: pd.DataFrame,
    templates: dict[str, bytes],
    logger: logging.Logger,
) -> tuple[list[tuple[str, bytes]], dict, list[dict], list[dict]]:
    """
    Run the pipeline for a single issuer.
    Returns (output_files, summary, audit_records, validation_records).
    Output files are ONLY the populated templates.
    """
    outputs: list[tuple[str, bytes]] = []
    all_template_headers: dict[str, dict[str, list]] = {}
    all_audit_records: list[dict] = []
    total_dvv_overrides = 0
    total_fields = 0
    errors: list[str] = []

    logger.info("=" * 60)
    logger.info(f"Processing issuer: {issuer_id}  ({len(issuer_ext_df)} rows)")

    for tmpl_key, tmpl_bytes in templates.items():
        logger.info(f"  -- {tmpl_key} --")

        try:
            wb, sheet_meta = populate_template(
                tmpl_bytes, tmpl_key, issuer_ext_df, issuer_id, logger
            )
        except Exception as e:
            msg = f"Population failed for '{tmpl_key}': {e}"
            logger.error(msg)
            logger.debug(traceback.format_exc())
            errors.append(msg)
            continue

        all_template_headers[tmpl_key] = {
            sn: m["headers"] for sn, m in sheet_meta.items()
        }
        for m in sheet_meta.values():
            total_fields += len(m["headers"]) * m["populated_rows"]

        try:
            audit_recs = apply_dvv_overrides(
                wb, sheet_meta, dvv_df, issuer_id, tmpl_key, logger
            )
            all_audit_records.extend(audit_recs)
            total_dvv_overrides += sum(
                1 for r in audit_recs if r["Status"].startswith("SUCCESS")
            )
        except Exception as e:
            msg = f"DVV override failed for '{tmpl_key}': {e}"
            logger.error(msg)
            logger.debug(traceback.format_exc())
            errors.append(msg)

        try:
            for sn in wb.sheetnames:
                if sn in sheet_meta:
                    m = sheet_meta[sn]
                    format_worksheet(
                        wb[sn], m["data_start_row"], m["populated_rows"], logger
                    )
        except Exception as e:
            logger.error(f"Formatting failed for '{tmpl_key}': {e}")
            logger.debug(traceback.format_exc())

        out_name = f"{issuer_id}_{CONFIG['OUTPUT_NAMES'][tmpl_key]}.xlsx"
        try:
            outputs.append((out_name, _wb_to_bytes(wb)))
            logger.info(f"  Built: {out_name}")
        except Exception as e:
            msg = f"Save failed for '{tmpl_key}': {e}"
            logger.error(msg)
            errors.append(msg)

    # Validation — computed for on-screen display only (no file written).
    val_records = build_validation_log(
        all_template_headers, issuer_ext_df, issuer_id, logger
    )

    summary = {
        "Issuer ID": issuer_id,
        "Records Processed": len(issuer_ext_df),
        "Fields Populated": total_fields,
        "DVV Overrides": total_dvv_overrides,
        "Validation Issues": len(val_records),
        "Errors": errors,
    }
    return outputs, summary, all_audit_records, val_records


def run_pipeline(
    csv_bytes: bytes, csv_name: str,
    dvv_bytes: bytes, dvv_name: str,
    templates: dict[str, bytes],
    logger: logging.Logger,
) -> dict:
    """
    Full pipeline. Detects issuer ID(s) from the CSV, processes each, and
    bundles the populated template files into a single ZIP.
    Returns a result dict.
    """
    logger.info("=" * 60)
    logger.info("STEP 1: Loading extraction CSV")
    ext_df = load_extraction(csv_bytes, csv_name, logger)

    logger.info("=" * 60)
    logger.info("STEP 2: Detecting issuer ID(s)")
    issuer_ids = detect_issuer_ids(ext_df, logger)

    logger.info("=" * 60)
    logger.info("STEP 3: Loading DVV file")
    dvv_df = load_dvv(dvv_bytes, dvv_name, logger)

    issuer_col = CONFIG["ISSUER_ID_FIELD"]
    # One issuer -> use the full CSV. Many issuers -> filter the CSV per issuer.
    if len(issuer_ids) == 1:
        jobs = [(issuer_ids[0], ext_df)]
    else:
        jobs = [
            (iid, ext_df[ext_df[issuer_col].astype(str).str.strip() == iid].copy())
            for iid in issuer_ids
        ]

    logger.info("=" * 60)
    logger.info("STEP 4: Populating templates and applying DVV overrides")

    all_outputs: list[tuple[str, bytes]] = []
    summaries: list[dict] = []
    audit_records: list[dict] = []
    validation_records: list[dict] = []

    for issuer_id, issuer_df in jobs:
        outs, summ, audits, vals = process_one_issuer(
            issuer_id, issuer_df, dvv_df, templates, logger
        )
        if len(jobs) > 1:
            outs = [(f"{issuer_id}/{name}", data) for name, data in outs]
        all_outputs.extend(outs)
        summaries.append(summ)
        audit_records.extend(audits)
        validation_records.extend(vals)

    logger.info("=" * 60)
    logger.info("STEP 5: Building ZIP archive (templates only)")
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in all_outputs:
            zf.writestr(name, data)
    zip_buf.seek(0)

    total_errors = sum(len(s["Errors"]) for s in summaries)
    logger.info("=" * 60)
    logger.info(f"DONE. Issuers: {len(summaries)} | "
                f"Files: {len(all_outputs)} | Errors: {total_errors}")

    return {
        "zip_bytes": zip_buf.getvalue(),
        "summaries": summaries,
        "n_files": len(all_outputs),
        "audit_records": audit_records,
        "validation_records": validation_records,
    }


# ==============================================================================
# SECTION 10: STREAMLIT UI
# ==============================================================================

def main() -> None:
    st.set_page_config(
        page_title="MSCI Annual Update Processor",
        page_icon="📄",
        layout="centered",
    )

    st.title("MSCI Annual Update Factual Process")
    st.caption(
        "Upload the extraction CSV and the DVV merged file, then run. "
        "The three templates are fixed (bundled with the app) and the Issuer ID "
        "is read automatically from the `DMX_ISSUER_ID` column of the CSV."
    )

    # ---- Load the fixed templates (bundled with the app) --------------------
    boot_logger, _ = setup_logging()
    fixed_templates, _missing = load_fixed_templates(boot_logger)

    # ---- Uploads ------------------------------------------------------------
    st.subheader("1 · Input files")
    csv_file = st.file_uploader(
        "Extraction CSV", type=["csv"], key="csv",
        help="The IID..._IssuerName.csv extraction file.",
    )
    dvv_file = st.file_uploader(
        "DVV merged file (XLSX)", type=["xlsx"], key="dvv",
        help="Columns are matched by header name: DATALIB_TAG (Data Lib), "
             "CORRECT_VALUE (Correct Value), SERIALID (Series ID).",
    )

    # ---- Fixed templates (loaded silently) ----------------------------------
    # The status panel is intentionally not shown. A single error is surfaced
    # only if no templates could be found at all, so the app never fails
    # silently.
    if not fixed_templates:
        st.error(
            "No template files were found in the repository. Expected: "
            + ", ".join(CONFIG["TEMPLATE_FILES"].values())
            + ". Commit them to the repo root (or a 'templates/' folder)."
        )

    # ---- Run ----------------------------------------------------------------
    st.subheader("2 · Run")
    ready = (
        csv_file is not None
        and dvv_file is not None
        and len(fixed_templates) > 0
    )
    if not ready and fixed_templates:
        st.info("Upload a CSV and a DVV file to enable Run.")

    run = st.button("▶ Run", type="primary", disabled=not ready, use_container_width=True)

    if run:
        logger, log_buffer = setup_logging()
        try:
            with st.spinner("Processing…"):
                result = run_pipeline(
                    csv_file.getvalue(), csv_file.name,
                    dvv_file.getvalue(), dvv_file.name,
                    fixed_templates, logger,
                )
            result["log_text"] = log_buffer.getvalue()
            st.session_state["result"] = result
            st.success("Processing complete.")
        except Exception as e:
            st.session_state.pop("result", None)
            st.error(f"Processing failed: {e}")
            with st.expander("Execution log", expanded=True):
                st.code(log_buffer.getvalue() or "(no log)", language="text")

    # ---- Results ------------------------------------------------------------
    result = st.session_state.get("result")
    if result:
        st.subheader("3 · Results")

        summaries = result.get("summaries", [])
        total_records = sum(s["Records Processed"] for s in summaries)
        total_dvv     = sum(s["DVV Overrides"] for s in summaries)
        total_val     = sum(s["Validation Issues"] for s in summaries)
        total_err     = sum(len(s["Errors"]) for s in summaries)

        m = st.columns(4)
        m[0].metric("Issuers", len(summaries))
        m[1].metric("Records", total_records)
        m[2].metric("DVV overrides", total_dvv)
        m[3].metric("Validation issues", total_val)

        if total_err:
            st.warning(f"{total_err} error(s) occurred — see the per-issuer detail and log.")

        if summaries:
            table = pd.DataFrame([{
                "Issuer ID": s["Issuer ID"],
                "Records": s["Records Processed"],
                "Fields populated": s["Fields Populated"],
                "DVV overrides": s["DVV Overrides"],
                "Validation issues": s["Validation Issues"],
                "Errors": len(s["Errors"]),
            } for s in summaries])
            st.dataframe(table, use_container_width=True, hide_index=True)

        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        if len(summaries) == 1:
            zip_name = f"{summaries[0]['Issuer ID']}_Output_{ts}.zip"
        else:
            zip_name = f"AnnualUpdate_Output_{ts}.zip"

        st.download_button(
            label=f"⬇ Download outputs ({result.get('n_files', 0)} files, ZIP)",
            data=result["zip_bytes"],
            file_name=zip_name,
            mime="application/zip",
            use_container_width=True,
        )

        # DVV audit trail (on-screen + optional CSV; not inside the ZIP).
        audit = result.get("audit_records", [])
        with st.expander(f"DVV audit trail ({len(audit)} record(s))"):
            if audit:
                adf = pd.DataFrame(audit)
                st.dataframe(adf, use_container_width=True, hide_index=True)
                st.download_button(
                    "⬇ DVV audit log (CSV)",
                    data=adf.to_csv(index=False).encode("utf-8"),
                    file_name="DVV_Audit_Log.csv",
                    mime="text/csv",
                )
            else:
                st.write("No DVV changes were applied.")

        # Validation issues (on-screen + optional CSV).
        vals = result.get("validation_records", [])
        with st.expander(f"Validation issues ({len(vals)})"):
            if vals:
                vdf = pd.DataFrame(vals)
                st.dataframe(vdf, use_container_width=True, hide_index=True)
                st.download_button(
                    "⬇ Validation issues (CSV)",
                    data=vdf.to_csv(index=False).encode("utf-8"),
                    file_name="Validation_Issues.csv",
                    mime="text/csv",
                )
            else:
                st.write("No validation issues found.")

        with st.expander("Execution log"):
            st.code(result.get("log_text", "") or "(no log)", language="text")


if __name__ == "__main__":
    main()
